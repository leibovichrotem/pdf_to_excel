from openpyxl import load_workbook, Workbook
import os, pandas as pd
import const


def is_transaction_exist(transaction_id):
    """
    This function check if transaction id allready exist in the excel file.
    if exist it's ignore duplicates.
    :param transaction_id: transaction id
    :return: boolean
    """
    xls = pd.ExcelFile(const.EXCEL_FILE_PATH)
    sheets = xls.sheet_names
    is_exist = False
    for sheet in sheets:
        if is_empty_sheet(sheet):
            pass
        else:
            df = xls.parse(sheet_name=sheet)
            sheet_dict = df.to_dict()
            transaction_list = list(sheet_dict['transaction_id'].values())
            if int(transaction_id) in transaction_list:
                is_exist = True
    return is_exist


def is_dir_or_file(sheet_name):
    """
    This function check if there is directory and file to excel file.
    if needed it creates dir/file according to const file.
    """
    if os.path.isdir(const.EXCEL_DIRECTORY_PATH) is False:
        print('//- Creating new directory and file for bank statements..')
        os.mkdir(const.EXCEL_DIRECTORY_PATH)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename=const.EXCEL_FILE_PATH)
    elif os.path.isfile(const.EXCEL_FILE_PATH) is False:
        print('//- Creating new excel file for bank statements..')
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename=const.EXCEL_FILE_PATH)


def is_sheet_exist(sheet_name):
    """
    This function chek if given sheet name is exist in the excel file.
    :param sheet_name: sheet name
    :return: boolean
    """
    xl = pd.ExcelFile(const.EXCEL_FILE_PATH)
    sheets = xl.sheet_names
    if sheet_name in sheets:
        return True
    else:
        return False


def is_empty_sheet(sheet_name):
    """
    This function check if sheet is empty.
    its important to know if we add data or write it from the start.
    """
    xl = pd.ExcelFile(const.EXCEL_FILE_PATH)
    is_empty = True
    df = xl.parse(sheet_name)
    sheet_dict = df.to_dict()
    if len(sheet_dict) > 0:
        is_empty = False
    return is_empty


def append_new_sheet(sheet_name):
    """
    This function create a new sheet in the excl file
    """
    print('//- Opening a new bank account sheet in excel file..')
    wb = load_workbook(const.EXCEL_FILE_PATH)
    sheet = wb.create_sheet(title=sheet_name)
    sheet.title = sheet_name
    wb.save(const.EXCEL_FILE_PATH)


def write_to_empty_sheet(sheet_name, statement):
    """
    This function write data to an empty sheet.
    it will write the columns of the data.
    :param sheet_name: sheet name
    :param statement: data to insert as dict
    """
    xl = pd.ExcelFile(const.EXCEL_FILE_PATH)
    xl_dict = xl.parse(sheet_name=None)
    xl_sheets = list(xl_dict.keys())
    df_list = []
    writer = pd.ExcelWriter(const.EXCEL_FILE_PATH, engine='openpyxl')
    for sheet in xl_sheets:
        df = xl_dict[sheet]
        if sheet_name == sheet:
            df = pd.DataFrame(statement)
            df_list.append({sheet: df})
        else:
            df_list.append({sheet: df})

    for df_dict in df_list:
        sheet = list(df_dict.keys())[0]
        df = df_dict[sheet]
        df = df.astype(str)
        df.to_excel(writer, sheet, index=False)
    print('//- Writing to new and empty sheet')
    writer.save()


def rewrite_excel(sheet_name, statement):
    """
    This function copy the excel file data, and insert the new data into it.
    Than its overwrite the old data + new data.
    """
    xl = pd.ExcelFile(const.EXCEL_FILE_PATH)
    xl_dict = xl.parse(sheet_name=None)
    xl_sheets = list(xl_dict.keys())
    df_list = []
    writer = pd.ExcelWriter(const.EXCEL_FILE_PATH, engine='openpyxl')
    for sheet in xl_sheets:
        if sheet == sheet_name:
            df = xl_dict[sheet]
            sheet_dict = df.to_dict()
            for column in sheet_dict:
                index = list(sheet_dict[column].keys())[-1]
                sheet_dict[column][index + 1] = statement[column][0]
            appended_df = pd.DataFrame(sheet_dict)
            df_list.append({sheet: appended_df})
        else:
            df = xl_dict[sheet]
            df_list.append({sheet: df})

    for df_dict in df_list:
        sheet = list(df_dict.keys())[0]
        df = df_dict[sheet]
        df = df.astype(str)
        df.to_excel(writer, sheet, index=False)
    print('//- Rewriting excel file')
    writer.save()


def manager(sheet_name, statement):
    """
    This function manage the progress of writing to excel.
    the options :
    1. write data to an empty sheet
    2. append new sheet to file
    2. append data to file and overwrite the file
    """
    is_dir_or_file(sheet_name)
    transaction_id = statement['transaction_id'][0]
    if is_sheet_exist(sheet_name):
        if is_empty_sheet(sheet_name) is False:
            if not is_transaction_exist(transaction_id):
                rewrite_excel(sheet_name, statement)
        else:
            write_to_empty_sheet(sheet_name, statement)
    else:
        append_new_sheet(sheet_name)
        write_to_empty_sheet(sheet_name, statement)


